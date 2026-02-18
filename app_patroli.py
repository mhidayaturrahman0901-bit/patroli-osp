import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError, ImageFilter
import imagehash
import io
import os
import re
import sqlite3
import zipfile
import hashlib
import requests
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import math

# =========================
# CONFIG UI
# =========================
st.set_page_config(page_title="Audit Foto Patroli PRO", layout="wide")
st.title("🕵️ AUDIT FOTO PATROLI (CENTER-AUDIT MODE)")
st.caption("Mendukung Embedded Excel + Link Google Docs. Fokus Audit: Area Tengah Foto.")

# =========================
# DATABASE
# =========================
DB_PATH = "audit_history.db"

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS history (
            sha256 TEXT PRIMARY KEY,
            phash  TEXT,
            source_type TEXT,
            source_file TEXT,
            sheet TEXT,
            location TEXT,
            cluster TEXT,
            segment TEXT,
            url TEXT,
            first_seen DATE
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_phash ON history(phash)")
    return conn

def db_lookup(conn, sha256_hex: str, phash_str: str):
    exact = conn.execute("SELECT source_file, sheet, location, cluster, segment, url, first_seen FROM history WHERE sha256=?", (sha256_hex,)).fetchone()
    ph = conn.execute("SELECT source_file, sheet, location, cluster, segment, url, first_seen FROM history WHERE phash=? LIMIT 1", (phash_str,)).fetchone()
    return exact, ph

def db_insert(conn, row: dict):
    conn.execute("""
        INSERT OR IGNORE INTO history
        (sha256, phash, source_type, source_file, sheet, location, cluster, segment, url, first_seen)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (row["sha256"], row["phash"], row["source_type"], row["source_file"], row["sheet"], row["location"], row["cluster"], row["segment"], row["url"], row["first_seen"]))

# =========================
# LOGIKA HASHING & CENTER CROP
# =========================
def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def compute_hashes_from_bytes(img_bytes: bytes):
    try:
        img_full = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        w, h = img_full.size
        
        # POTONG TENGAH: Buang 25% Atas (Logo) & 25% Bawah (GEO Teks)
        top, bottom = int(h * 0.25), int(h * 0.75)
        left, right = int(w * 0.10), int(w * 0.90)
        audit_area = img_full.crop((left, top, right, bottom))

        thumb = audit_area.copy()
        thumb.thumbnail((240, 240))
        
        ph = str(imagehash.phash(thumb)) # Hash hanya dari area tengah
        sh = sha256_bytes(img_bytes)
        
        return sh, ph, thumb, img_full
    except:
        return None, None, None, None

# =========================
# DETEKSI LOGO & GEO (SMART FILTER)
# =========================
def image_entropy_gray(img: Image.Image) -> float:
    g = img.convert("L").resize((256, 256))
    hist = g.histogram()
    total = sum(hist)
    if total == 0: return 0.0
    ent = 0.0
    for h in hist:
        if h:
            p = h / total
            ent -= p * math.log2(p)
    return ent

def edge_density(img: Image.Image) -> float:
    g = img.convert("L").resize((256, 256)).filter(ImageFilter.FIND_EDGES)
    px = list(g.getdata())
    return (sum(px) / (len(px) or 1)) / 255.0

def classify_for_audit(img: Image.Image) -> tuple[bool, str, str]:
    # Jika gambar terlalu flat/polos (Logo-Only)
    ent = image_entropy_gray(img)
    ed = edge_density(img)
    if ent < 3.2 and ed < 0.035:
        return False, "⏭️ SKIP (Logo-Only)", f"Entropi rendah: {ent:.2f}"
    return True, "", ""

# =========================
# GOOGLE DOCS & LINK HANDLER
# =========================
DOC_ID_RE = re.compile(r"/document/d/([a-zA-Z0-9_-]+)")

def download_images_from_url(url: str) -> list[bytes]:
    if not url or not isinstance(url, str): return []
    m = DOC_ID_RE.search(url)
    if m:
        doc_id = m.group(1)
        export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=docx"
        try:
            r = requests.get(export_url, timeout=20)
            if r.status_code == 200:
                out = []
                with zipfile.ZipFile(io.BytesIO(r.content), "r") as z:
                    for name in z.namelist():
                        if name.startswith("word/media/"):
                            out.append(z.read(name))
                return out
        except: pass
    return []

# =========================
# EXCEL & AUDIT ENGINE
# =========================
def find_header_row_and_cols(ws):
    # Logika mencari kolom Cluster, Segment, Link
    col_cluster = col_segment = col_link = 1
    for r in range(1, 15):
        for c in range(1, 15):
            v = str(ws.cell(r, c).value).lower() if ws.cell(r, c).value else ""
            if "cluster" in v: col_cluster = c
            if "segment" in v: col_segment = c
            if "link" in v or "url" in v: col_link = c
    return r, col_cluster, col_segment, col_link

def audit_workbook(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    conn = get_db()
    all_results = []

    for ws in wb.worksheets:
        h_row, c_clu, c_seg, c_lin = find_header_row_and_cols(ws)
        
        # Proses Link Google Docs
        for r in range(h_row + 1, ws.max_row + 1):
            url = str(ws.cell(r, c_lin).value) if ws.cell(r, c_lin).value else ""
            if "docs.google.com" in url:
                img_list = download_images_from_url(url)
                for i, img_bytes in enumerate(img_list):
                    sh, ph, thumb, full = compute_hashes_from_bytes(img_bytes)
                    if full:
                        ok, skip_status, reason = classify_for_audit(full)
                        if not ok: continue
                        
                        exact, sim = db_lookup(conn, sh, ph)
                        status = "✅ VALID"
                        if exact: status = "❌ GUGUR (Exact)"
                        elif sim: status = "⚠️ CEK MANUAL (Mirip)"
                        
                        if status == "✅ VALID":
                            db_insert(conn, {"sha256": sh, "phash": ph, "source_type": "GDocs", "source_file": os.path.basename(xlsx_path), "sheet": ws.title, "location": f"R{r}", "cluster": str(ws.cell(r, c_clu).value), "segment": str(ws.cell(r, c_seg).value), "url": url, "first_seen": datetime.now().strftime("%Y-%m-%d")})
                        
                        all_results.append({"Sheet": ws.title, "Baris": r, "Status": status, "Thumb": thumb})
    
    conn.commit()
    return pd.DataFrame(all_results)

# =========================
# RUN STREAMLIT
# =========================
uploaded = st.file_uploader("Upload Excel Patroli", type=["xlsx"])
if uploaded:
    with open("temp.xlsx", "wb") as f: f.write(uploaded.getbuffer())
    if st.button("🚀 MULAI AUDIT"):
        df_res = audit_workbook("temp.xlsx")
        if not df_res.empty:
            st.success(f"Audit Selesai! Menemukan {len(df_res)} foto.")
            st.dataframe(df_res.drop(columns=["Thumb"]))
            
            # Preview Galeri
            cols = st.columns(4)
            for idx, row in df_res.iterrows():
                with cols[idx % 4]:
                    st.image(row["Thumb"], caption=f"Baris {row['Baris']}: {row['Status']}")
        else:
            st.warning("Tidak ada foto yang ditemukan atau semua link docs tidak bisa diakses.")
